import openpyxl
from openpyxl.styles import Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
#引きたい線の情報をここで設定
side = Side(style = "thin", color = "000000")
border = Border(bottom=side)

#ブックとシートの読み込み
wb = openpyxl.load_workbook("test.xlsx")
sheet = wb["Sheet1"]

#セルを指定して色情報を取得
color = sheet["C3"].fill.start_color.rgb
row = 23

num = 10#検索結果件数

list1 = ["A23:D23","E23:F23","G23:J23", "K23:L23", "M23:P23", "Q23:R23"]#色付きの範囲
list2 = ["A23", "E23", "G23", "K23", "M23", "Q23"]#色なしの方
#--------------------------------------------------------------------------
#列の高さを調節
for i in range(num):
    sheet.row_dimensions[row + i].height = 31.5
        
#cellのマージと色付け
for count in range(num):
    for i in range(len(list1)):
        index_1 = list1[i].replace(str(row), str(row + count))
        index_2 = list2[i].replace(str(row), str(row + count))
        print(index_1)
        sheet.merge_cells(index_1)
        if i % 2 == 0:
            fill = PatternFill(start_color=color, end_color=color, fill_type = "solid")
            sheet[index_2].fill = fill
            
#実際に書き込む
loop_counter = 0
for count in range(num):
    sheet[list2[0].replace(str(row), str(row + loop_counter))] = "あ"
    sheet[list2[1].replace(str(row), str(row + loop_counter))] = "い"
    sheet[list2[2].replace(str(row), str(row + loop_counter))] = "う"
    sheet[list2[3].replace(str(row), str(row + loop_counter))] = "え"
    sheet[list2[4].replace(str(row), str(row + loop_counter))] = "お"
    sheet[list2[5].replace(str(row), str(row + loop_counter))] = "か"
    
    #文字をセンター寄せにしている
    sheet[list2[0].replace(str(row), str(row + loop_counter))].alignment = Alignment(horizontal="center", vertical="center")
    sheet[list2[1].replace(str(row), str(row + loop_counter))].alignment = Alignment(horizontal="center", vertical="center")
    sheet[list2[2].replace(str(row), str(row + loop_counter))].alignment = Alignment(horizontal="center", vertical="center")
    sheet[list2[3].replace(str(row), str(row + loop_counter))].alignment = Alignment(horizontal="center", vertical="center")
    sheet[list2[4].replace(str(row), str(row + loop_counter))].alignment = Alignment(horizontal="center", vertical="center")
    sheet[list2[5].replace(str(row), str(row + loop_counter))].alignment = Alignment(horizontal="center", vertical="center")

    loop_counter += 1

#------------------------------------------------------------------------------------------   

#この操作は既に書き込まれている最終列行を取り出せる．
print(sheet.max_row) #行
print(sheet.max_column) #Rまで
print(get_column_letter(1))

#最終行に線を引いている 
for cell in sheet[sheet.max_row]:
    cell.border = border
wb.save("sample3.xlsx")
wb.close()
